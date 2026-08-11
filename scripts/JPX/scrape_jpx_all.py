#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JPX 一括スクレイパー  v1.6
  1. ToSTNeT 超大口約定情報       → GID 1044963009
  2. 自己株式立会外買付取引情報   → GID 2116229549
  3. 業種別 33 業種指数           → GID 742855575  ※Playwright 使用

TARGET_DATE 環境変数 (YYYYMMDD):
  - 指定あり → その日付「以降」（未来日含む）のデータを取得
  - 指定なし → ページ上の最新日を自動取得
"""

import io
import json
import logging
import os
import re
import time
from datetime import datetime, timezone, timedelta

import requests
from bs4 import BeautifulSoup
import gspread

SPREADSHEET_ID = "1kWST0CkkIvo3irPSbMgtVtUqqRDXFwvRREYZDRQAFMY"
GID_TOSTNET    = 1044963009
GID_OWN_SHARES = 2116229549
GID_SECTOR     = 742855575

TOSTNET_URL    = "https://www.jpx.co.jp/markets/equities/tostnet/index.html"
OWN_SHARES_URL = "https://www.jpx.co.jp/markets/equities/off-auction-ownshares/index.html"
SECTOR_URL     = "https://www.jpx.co.jp/markets/indices/realvalues/"
JPX_BASE       = "https://www.jpx.co.jp"

SECTOR_LABELS = {"current": "現在値", "low": "安値", "high": "高値", "open": "始値"}
SECTOR_BLOCK_ORDER = ["current", "low", "high", "open"]
SECTOR_GAP = 3   # ブロック間の空白行数

def _compute_sector_rows(n_sectors: int) -> dict:
    """
    各ブロックの開始行を動的に計算する。
    ブロック構成: 1行目=取得日(C列)、2〜n+1行目=データ(A列名/C列値)
    ブロック間は SECTOR_GAP 行だけ空ける。
    """
    block_size = 1 + n_sectors   # 日付行 + データ行
    step = block_size + SECTOR_GAP
    return {name: 1 + i * step for i, name in enumerate(SECTOR_BLOCK_ORDER)}

# マーカーは 1 番目のブロック(現在値)の C1（取得日）を流用
SECTOR_MARKER = "C1"

# 業種別指数ページの実測列順（Playwright レンダリング後）:
#   0=指数名 1=現在値 2=前日比(値) 3=前日比(%) 4=始値 5=高値 6=安値
SECTOR_COL = {"name": 0, "current": 1, "open": 4, "high": 5, "low": 6}

# 東証33業種
SECTOR33 = [
    "水産・農林業", "鉱業", "建設業", "食料品", "繊維製品", "パルプ・紙",
    "化学", "医薬品", "石油・石炭製品", "ゴム製品", "ガラス・土石製品",
    "鉄鋼", "非鉄金属", "金属製品", "機械", "電気機器", "輸送用機器",
    "精密機器", "その他製品", "電気・ガス業", "陸運業", "海運業", "空運業",
    "倉庫・運輸関連業", "情報・通信業", "卸売業", "小売業", "銀行業",
    "証券・商品先物取引業", "保険業", "その他金融業", "不動産業", "サービス業",
]

def _norm(s: str) -> str:
    s = s.replace("\u3000", "").replace(" ", "").replace("　", "")
    s = s.replace("・", "").replace("･", "")
    s = s.replace("（", "(").replace("）", ")")
    return s.strip()

SECTOR33_NORM = {_norm(x): x for x in SECTOR33}

SKIP_NAMES = frozenset([
    "指数名", "銘柄", "業種", "指数", "現在値", "前日比", "始値", "高値", "安値",
    "前日終値", "騰落率", "リアルタイムグラフ", "ヒストリカルグラフ",
    "データ更新日時", "", "-", "－", "−",
])

JST      = timezone(timedelta(hours=9))
DATE_PAT = re.compile(r"^\d{4}[/\-年]\d{1,2}[/\-月]\d{1,2}")

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
#  日付ユーティリティ
# ─────────────────────────────────────────────────────────────
def parse_jp_date(s: str):
    """'2026/08/10' 等を date オブジェクトに変換。失敗時 None。"""
    s = s.strip()
    m = re.match(r"(\d{4})[/\-年](\d{1,2})[/\-月](\d{1,2})", s)
    if not m:
        return None
    y, mo, d = map(int, m.groups())
    try:
        return datetime(y, mo, d).date()
    except ValueError:
        return None


def resolve_target_date():
    """TARGET_DATE 環境変数 (YYYYMMDD) → 'YYYY/MM/DD' 文字列 or None"""
    raw = os.environ.get("TARGET_DATE", "").strip()
    if raw and re.fullmatch(r"\d{8}", raw):
        d = f"{raw[:4]}/{raw[4:6]}/{raw[6:8]}"
        log.info("対象日: %s（指定・以降を取得）", d)
        return d
    log.info("対象日: 未指定 → 各ページの最新日を自動取得")
    return None


# ─────────────────────────────────────────────────────────────
#  Google Sheets
# ─────────────────────────────────────────────────────────────
def build_gc():
    info = json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"])
    return gspread.service_account_from_dict(info)

def open_ss(gc):
    return gc.open_by_key(SPREADSHEET_ID)

def get_ws(ss, gid):
    for ws in ss.worksheets():
        if ws.id == gid:
            return ws
    raise ValueError(f"GID {gid} not found")

def prepend_rows(ws, rows):
    ws.insert_rows(rows, row=2, value_input_option="USER_ENTERED")
    log.info("    ✓ %d 行を row=2 に挿入", len(rows))

def insert_col_c(ss, ws):
    ss.batch_update({"requests": [{"insertDimension": {
        "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                  "startIndex": 2, "endIndex": 3},
        "inheritFromBefore": False,
    }}]})
    log.info("    ✓ 列 C に新列挿入")


# ─────────────────────────────────────────────────────────────
#  HTTP
# ─────────────────────────────────────────────────────────────
def fetch(url):
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or "utf-8"
    return BeautifulSoup(resp.text, "html.parser")


# ═════════════════════════════════════════════════════════════
#  1. ToSTNeT 超大口約定情報
#  リンク先は xlsx。ヘッダー行を検出して動的に列マッピングする。
#  出力列: 公表日 | 取引日 | コード | 銘柄 | 価格 | 売買高 | 売買代金
# ═════════════════════════════════════════════════════════════
# 実際の xlsx ヘッダー（1行目）:
#   公表日/Publication_Date, 取引日/Trading_Date, 約定時刻/Trade_Time,
#   銘柄コード/Code, 銘柄名_日本語/Issue_Name_Japanese, 銘柄名_英語/Issue_Name_English,
#   価格_円/Price_yen, 売買高_株/Trading_Volume_shares, 売買代金_円/Trading_Value_yen
_TOSTNET_HDR_KEYS = {
    "kouhyo":  ["公表日", "Publication_Date"],
    "date":    ["取引日", "Trading_Date"],
    "code":    ["銘柄コード", "Code"],
    "name":    ["銘柄名_日本語", "Issue_Name_Japanese"],
    "price":   ["価格", "Price"],
    "volume":  ["売買高", "Trading_Volume"],
    "value":   ["売買代金", "Trading_Value"],
}

# 実測ヘッダー順（0-indexed）のフォールバック位置:
#   0=公表日 1=取引日 2=約定時刻 3=銘柄コード 4=銘柄名_日本語
#   5=銘柄名_英語 6=価格_円 7=売買高_株 8=売買代金_円
_TOSTNET_FALLBACK_COLS = {
    "kouhyo": 0, "date": 1, "code": 3, "name": 4,
    "price": 6, "volume": 7, "value": 8,
}


def _detect_xlsx_header(rows):
    """
    2次元セル配列からヘッダー行を検出し、列インデックスマップを返す。
    見つからない場合は (None, {}) を返す。
    """
    for i, row in enumerate(rows[:15]):
        cells = ["" if c is None else str(c).strip() for c in row]
        joined = "".join(cells)
        if ("銘柄コード" in joined or "Code" in joined) and \
           ("価格" in joined or "Price" in joined):
            col = {}
            for j, c in enumerate(cells):
                for key, keywords in _TOSTNET_HDR_KEYS.items():
                    if key in col:
                        continue
                    if any(kw in c for kw in keywords):
                        col[key] = j
            if "code" in col:
                return i, col
    return None, {}


def _parse_tostnet_xlsx(url, kouhyo_date):
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=60)
    resp.raise_for_status()

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    out = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        log.info("  [xlsx] シート '%s' (%d行 x %d列)",
                 ws.title, len(rows), ws.max_column)

        header_idx, col = _detect_xlsx_header(rows)

        if header_idx is None:
            log.info("  [xlsx] ヘッダー未検出 → 先頭6行をダンプ、固定位置で継続")
            for j, row in enumerate(rows[:6]):
                cells = ["" if c is None else str(c) for c in row]
                log.info("    row[%d]: %s", j, [c[:20] for c in cells])
            header_idx = 0
            col = dict(_TOSTNET_FALLBACK_COLS)

        # 検出できなかったキーは既知の固定位置で補完
        for k, v in _TOSTNET_FALLBACK_COLS.items():
            col.setdefault(k, v)

        log.info("  [xlsx] ヘッダー行=%d 列マップ=%s", header_idx, col)

        for row in rows[header_idx + 1:]:
            cells = ["" if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue

            code_raw = cells[col["code"]] if "code" in col and col["code"] < len(cells) else ""
            code = code_raw.split(".")[0]
            if not re.fullmatch(r"\d{4}", code):
                continue

            def g(key, default=""):
                idx = col.get(key)
                if idx is None or idx >= len(cells):
                    return default
                v = cells[idx]
                return v

            def g_num(key, default=""):
                """数値セル用: openpyxl が float(1234.0) を返すケースを整形"""
                v = g(key, default)
                if re.fullmatch(r"-?\d+\.0", v):
                    v = v[:-2]
                return v

            # 公表日はシート内の値を優先、無ければ引数の kouhyo_date を使用
            kouhyo = g("kouhyo", kouhyo_date) or kouhyo_date
            torihiki = g("date")
            name     = g("name")
            price    = g_num("price")
            volume   = g_num("volume")
            value    = g_num("value")

            out.append([kouhyo, torihiki, code, name, price, volume, value])

    return out


def _parse_tostnet_html(soup, kouhyo_date):
    """リンク先が HTML の場合のフォールバック"""
    out = []
    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all("td")]
            if not cells:
                continue
            for idx, c in enumerate(cells):
                if re.fullmatch(r"\d{4}", c) and idx >= 1:
                    row = [kouhyo_date] + cells[max(0, idx - 1): idx + 6]
                    if len(row) >= 7:
                        out.append(row[:7])
                    break
    return out


def scrape_tostnet(target_date=None):
    """
    target_date (YYYY/MM/DD): この日付「以降」の公表日を対象にリンクを追う
    None: テーブル最上行（最新日）のみ
    """
    soup = fetch(TOSTNET_URL)
    out = []
    target_d = parse_jp_date(target_date) if target_date else None

    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            tds = tr.find_all("td")
            if not tds:
                continue
            kouhyo = tds[0].get_text(strip=True)
            if not DATE_PAT.match(kouhyo):
                continue

            kouhyo_d = parse_jp_date(kouhyo)
            if target_d and kouhyo_d and kouhyo_d < target_d:
                continue   # 指定日より過去はスキップ（未来日は含む）
            if not target_d:
                pass  # 最新行のみ後段の break で担保

            link_tag = tds[1].find("a", href=True) if len(tds) > 1 else None
            if not link_tag:
                if not target_d:
                    break
                continue

            href = link_tag["href"]
            if not href.startswith("http"):
                href = JPX_BASE + href

            log.info("  ToSTNeT: %s → %s", kouhyo, href)
            try:
                if href.lower().endswith((".xlsx", ".xls")):
                    rows = _parse_tostnet_xlsx(href, kouhyo)
                else:
                    rows = _parse_tostnet_html(fetch(href), kouhyo)
                log.info("  → %d 件取得", len(rows))
                out.extend(rows)
            except Exception as e:
                log.warning("  取得失敗: %s", e, exc_info=True)

            if not target_d:
                break   # 最新日のみ

    return out


def run_tostnet(ss, target_date):
    log.info("  スクレイプ: ToSTNeT 超大口...")
    rows = scrape_tostnet(target_date)
    log.info("  → %d 件", len(rows))
    if not rows:
        log.warning("  データなし → スキップ")
        return
    ws = get_ws(ss, GID_TOSTNET)
    cur = ws.row_values(2)
    if cur and rows and cur[0] == rows[0][0] and not target_date:
        log.info("  既挿入済み (%s) → スキップ", rows[0][0])
        return
    prepend_rows(ws, rows)


# ═════════════════════════════════════════════════════════════
#  2. 自己株式立会外買付取引情報
#  実施日が target_date「以降」（未来日含む）の全行を取得
#  出力列: 公表日(今日) | 実施日 | コード | 銘柄 | 価格 | 買付数量 | 約定数量
# ═════════════════════════════════════════════════════════════
def _parse_name_code(text):
    m = re.search(r"[（(](\d{4})[）)]", text)
    code = m.group(1) if m else ""
    name = text[:m.start()].strip() if m else text
    name = re.sub(r"[\u3000\s]+(株式|株|全株)\s*$", "", name).strip()
    return name, code


def scrape_own_shares(target_date=None):
    soup = fetch(OWN_SHARES_URL)
    today = datetime.now(JST).strftime("%Y/%m/%d")
    target_d = parse_jp_date(target_date) if target_date else None
    out = []

    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue
            jisshi = tds[0].get_text(strip=True)
            if not DATE_PAT.match(jisshi):
                continue

            if target_d:
                jisshi_d = parse_jp_date(jisshi)
                if jisshi_d and jisshi_d < target_d:
                    continue   # 指定日より過去はスキップ（未来日は含む）

            name_code = tds[1].get_text(strip=True)
            price     = tds[2].get_text(strip=True)
            buy_qty   = tds[3].get_text(strip=True)
            exec_qty  = tds[4].get_text(strip=True) if len(tds) > 4 else ""

            name, code = _parse_name_code(name_code)
            out.append([today, jisshi, code, name, price, buy_qty, exec_qty])

    return out


def run_own_shares(ss, target_date):
    log.info("  スクレイプ: 自己株式立会外買付...")
    rows = scrape_own_shares(target_date)
    log.info("  → %d 件", len(rows))
    if not rows:
        log.warning("  データなし → スキップ")
        return
    ws = get_ws(ss, GID_OWN_SHARES)
    cur = ws.row_values(2)
    if cur and len(cur) > 1 and rows and cur[1] == rows[0][1] and not target_date:
        log.info("  既挿入済み (実施日: %s) → スキップ", rows[0][1])
        return
    prepend_rows(ws, rows)


# ═════════════════════════════════════════════════════════════
#  3. 業種別 33 業種指数  ※JS レンダリング → Playwright 使用
#  実測列順: 0=指数名 1=現在値 2=前日比(値) 3=前日比(%) 4=始値 5=高値 6=安値
# ═════════════════════════════════════════════════════════════
def _parse_sector_html(soup):
    max_idx = max(SECTOR_COL.values())
    out, seen = [], set()

    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
            if len(cells) <= max_idx:
                continue

            name = cells[SECTOR_COL["name"]]
            if not name or name in SKIP_NAMES:
                continue

            cur = cells[SECTOR_COL["current"]]
            if not cur or cur in SKIP_NAMES:
                continue
            if not re.search(r"[\d０-９,，.]", cur):
                continue

            key = _norm(name)
            if key not in SECTOR33_NORM:
                continue
            canonical = SECTOR33_NORM[key]

            if canonical in seen:
                continue
            seen.add(canonical)

            out.append({
                "name":    canonical,
                "current": cur,
                "high":    cells[SECTOR_COL["high"]],
                "low":     cells[SECTOR_COL["low"]],
                "open":    cells[SECTOR_COL["open"]],
            })

    order = {n: i for i, n in enumerate(SECTOR33)}
    out.sort(key=lambda d: order.get(d["name"], 999))
    return out


def _fetch_with_playwright():
    from playwright.sync_api import sync_playwright
    log.info("  Playwright 起動中...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
        page = browser.new_page()
        page.goto(SECTOR_URL, wait_until="networkidle", timeout=30_000)
        try:
            page.wait_for_function(
                "document.querySelectorAll('table tr').length > 14",
                timeout=10_000,
            )
        except Exception:
            log.info("  待機タイムアウト（現状の HTML を使用）")
        html = page.content()
        browser.close()
    log.info("  Playwright: HTML 取得完了")
    return html


def scrape_sector():
    soup = fetch(SECTOR_URL)
    data = _parse_sector_html(soup)
    if not data:
        log.info("  requests: 0件 → Playwright で再試行...")
        try:
            html = _fetch_with_playwright()
            soup2 = BeautifulSoup(html, "html.parser")
            data = _parse_sector_html(soup2)
            log.info("  Playwright後: %d 業種", len(data))

            if data:
                # 検証用: 最初の1件の生セルをログ出力（列ズレ確認用）
                for tbl in soup2.find_all("table"):
                    for tr in tbl.find_all("tr"):
                        cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
                        if len(cells) > max(SECTOR_COL.values()) and cells[0] == data[0]["name"]:
                            log.info("  [検証] %s の生データ: %s", data[0]["name"], cells[:8])
                            break
                    else:
                        continue
                    break
            else:
                tables = soup2.find_all("table")
                log.info("  [DEBUG Playwright後] テーブル数=%d", len(tables))
                for i, tbl in enumerate(tables[:4]):
                    trs = tbl.find_all("tr")
                    for j, tr in enumerate(trs[:3]):
                        cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
                        log.info("    table[%d] row[%d] %d列: %s",
                                 i, j, len(cells), [c[:20] for c in cells])
        except ImportError:
            log.error("  Playwright 未インストール")
        except Exception as e:
            log.error("  Playwright エラー: %s", e, exc_info=True)
    return data


def write_sector_block(ws, data, metric, start_row, fetch_date):
    """
    start_row     : 取得日を書く行（C列）
    start_row + 1 〜 start_row + n : 業種名(A列) と 値(C列)
    """
    n = len(data)
    if n == 0:
        return

    # 1行目: 取得日 (C列)
    ws.update(values=[[fetch_date]], range_name=f"C{start_row}",
              value_input_option="USER_ENTERED")

    # 2行目〜: 業種名(A) と 値(C)
    data_start = start_row + 1
    ws.update(values=[[d["name"]] for d in data],
              range_name=f"A{data_start}:A{data_start+n-1}",
              value_input_option="USER_ENTERED")
    ws.update(values=[[d[metric]] for d in data],
              range_name=f"C{data_start}:C{data_start+n-1}",
              value_input_option="USER_ENTERED")
    log.info("    ✓ %s: 日付行=%d データ行%d〜%d (%d件)",
             SECTOR_LABELS[metric], start_row, data_start, data_start+n-1, n)
    time.sleep(1.2)


def run_sector(ss):
    log.info("  スクレイプ: 業種別 33 業種指数...")
    data = scrape_sector()
    log.info("  → %d 業種取得", len(data))
    if not data:
        log.warning("  データなし → スキップ")
        return

    ws = get_ws(ss, GID_SECTOR)
    now_str = datetime.now(JST).strftime("%Y/%m/%d %H:%M")
    today = datetime.now(JST).strftime("%Y/%m/%d")

    # 重複チェック: C1（現在値ブロックの取得日）の日付部分を比較
    marker = (ws.acell(SECTOR_MARKER).value or "").strip()
    if marker[:10] == today:
        log.info("  既挿入済み (%s) → スキップ", today)
        return

    insert_col_c(ss, ws)
    time.sleep(1.5)

    sector_rows = _compute_sector_rows(len(data))
    for metric, start_row in sector_rows.items():
        write_sector_block(ws, data, metric, start_row, now_str)

    log.info("  ✓ 全ブロック書き込み完了 (取得日時: %s)", now_str)


# ═════════════════════════════════════════════════════════════
#  メイン
# ═════════════════════════════════════════════════════════════
def main():
    now = datetime.now(JST).strftime("%Y-%m-%d %H:%M JST")
    log.info("━━━ JPX 一括スクレイプ 開始: %s ━━━", now)

    target_date = resolve_target_date()

    gc = build_gc()
    ss = open_ss(gc)

    tasks = [
        ("ToSTNeT 超大口",     lambda: run_tostnet(ss, target_date)),
        ("自己株式立会外買付", lambda: run_own_shares(ss, target_date)),
        ("業種別 33 業種指数", lambda: run_sector(ss)),
    ]
    for label, fn in tasks:
        log.info("▶ %s", label)
        try:
            fn()
        except Exception as e:
            log.error("  ✗ エラー: %s", e, exc_info=True)
        time.sleep(2)

    log.info("━━━ 完了 ━━━")


if __name__ == "__main__":
    main()
